from PyQt5.QtWidgets import QApplication, QDialog,QGridLayout, QComboBox, QDateEdit,QFileDialog, QMessageBox, QWidget, QVBoxLayout, QLineEdit, QPushButton, QLabel, QMainWindow, QTableWidget, QTableWidgetItem
from PyQt5.QtCore import QDate
from openpyxl import Workbook
from datetime import datetime
import sqlite3


def connect_db():
    try:
        conn = sqlite3.connect('schedule.db')
        return conn
    except sqlite3.Error as e:
        print(f"Ошибка подключения к базе данных: {e}")
        return None

class MainWindow(QMainWindow):
    def __init__(self, user, login_window):
        super().__init__()
        self.user = user
        self.login_window = login_window
        self.setWindowTitle('Главное меню')
        self.setGeometry(600, 200, 800, 600)

        self.layout = QVBoxLayout()

        self.label = QLabel(f'Добро пожаловать, {user["name"]}!')
        self.layout.addWidget(self.label)

        self.schedule_button = QPushButton('Расписание на неделю')
        self.schedule_button.clicked.connect(self.show_schedule)
        self.layout.addWidget(self.schedule_button)

        self.export_button = QPushButton('Экспорт расписания в Excel')
        self.export_button.clicked.connect(self.export_schedule_to_excel)
        self.layout.addWidget(self.export_button)

        self.search_button = QPushButton('Найти расписание по дате')
        self.search_button.clicked.connect(self.search_schedule_by_date)
        self.layout.addWidget(self.search_button)

        self.logout_button = QPushButton('Выйти из аккаунта')
        self.logout_button.clicked.connect(self.logout)
        self.layout.addWidget(self.logout_button)

        main_widget = QWidget(self)
        main_widget.setLayout(self.layout)
        self.setCentralWidget(main_widget)

    def show_schedule(self):
        """Функция для отображения расписания на неделю."""
        conn = connect_db()
        if conn is None:
            print("Не удалось подключение к базе данных.")
            return

        cursor = conn.cursor()

        try:
            if self.user['role'] == 'преподаватель':
                query = '''
                    SELECT 
                        schedule.day, 
                        bell_schedule.start_time, 
                        bell_schedule.end_time, 
                        subjects.subject_name, 
                        classrooms.classroom_name, 
                        groups.group_name
                    FROM schedule
                    JOIN bell_schedule ON schedule.pair_number = bell_schedule.pair_number
                    JOIN subjects ON schedule.subject_id = subjects.id
                    JOIN classrooms ON schedule.classroom_id = classrooms.id
                    JOIN groups ON schedule.group_id = groups.id
                    WHERE schedule.teacher_id = ?
                    ORDER BY schedule.day, schedule.pair_number;
                '''
                cursor.execute(query, (self.user['teacher_id'],))
            else:
                query = '''
                    SELECT 
                        schedule.day, 
                        bell_schedule.start_time, 
                        bell_schedule.end_time, 
                        subjects.subject_name, 
                        teachers.name, 
                        classrooms.classroom_name
                    FROM schedule
                    JOIN bell_schedule ON schedule.pair_number = bell_schedule.pair_number
                    JOIN subjects ON schedule.subject_id = subjects.id
                    JOIN teachers ON schedule.teacher_id = teachers.id
                    JOIN classrooms ON schedule.classroom_id = classrooms.id
                    WHERE schedule.group_id = ?
                    ORDER BY schedule.day, schedule.pair_number;
                '''
                cursor.execute(query, (self.user['student_id'],))

            rows = cursor.fetchall()

            if not rows:
                QMessageBox.warning(self, "Ошибка", "Нет расписания.")
                return

            # Очистка виджета
            for i in reversed(range(self.layout.count())):
                widget = self.layout.itemAt(i).widget()
                if widget is not None:
                    widget.deleteLater()

            # Группировка расписания по дням
            schedule_by_day = {}
            for row in rows:
                day = row[0]
                if day not in schedule_by_day:
                    schedule_by_day[day] = []
                schedule_by_day[day].append(row[1:])

            days_order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]

            # Вывод расписания по дням
            for day in days_order:
                if day in schedule_by_day:
                    day_label = QLabel(f"<b>{day}</b>")
                    self.layout.addWidget(day_label)

                    day_table = QTableWidget()
                    entries = schedule_by_day[day]
                    columns = ['Начало', 'Конец', 'Предмет', 'Кабинет', 'Группа'] if self.user['role'] == 'преподаватель' else \
                              ['Начало', 'Конец', 'Предмет', 'Преподаватель', 'Кабинет']
                    day_table.setRowCount(len(entries))
                    day_table.setColumnCount(len(columns))
                    day_table.setHorizontalHeaderLabels(columns)

                    for i, entry in enumerate(entries):
                        for j, value in enumerate(entry):
                            day_table.setItem(i, j, QTableWidgetItem(str(value)))

                    day_table.resizeColumnsToContents()
                    day_table.resizeRowsToContents()

                    self.layout.addWidget(day_table)

            self.back_button = QPushButton('Вернуться в главное меню')
            self.back_button.clicked.connect(self.back_to_main_menu)
            self.layout.addWidget(self.back_button)

        except sqlite3.Error as e:
            print(f"Database query error: {e}")
        except Exception as e:
            print(f"Unexpected error: {e}")
        finally:
            conn.close()

    def export_schedule_to_excel(self):
        """Функция для экспорта расписания на неделю в формат Excel для преподавателя и студента."""
        conn = connect_db()
        if conn is None:
            print("Database connection failed.")
            return

        cursor = conn.cursor()

        try:
            if self.user['role'] == 'студент':
                cursor.execute('SELECT group_id FROM students WHERE id = ?', (self.user['student_id'],))
                group_result = cursor.fetchone()

                if not group_result:
                    self.label.setText("Error: Student is not assigned to any group.")
                    return

                group_id = group_result[0]

                query = '''
                    SELECT 
                        schedule.day, 
                        schedule.date, 
                        bell_schedule.start_time, 
                        bell_schedule.end_time, 
                        subjects.subject_name, 
                        teachers.name, 
                        classrooms.classroom_name
                    FROM schedule
                    JOIN bell_schedule ON schedule.pair_number = bell_schedule.pair_number
                    JOIN subjects ON schedule.subject_id = subjects.id
                    JOIN teachers ON schedule.teacher_id = teachers.id
                    JOIN classrooms ON schedule.classroom_id = classrooms.id
                    WHERE schedule.group_id = ?
                    ORDER BY schedule.day, schedule.pair_number;
                '''
                cursor.execute(query, (group_id,))
                rows = cursor.fetchall()

                if not rows:
                    self.label.setText("No schedule found for your group.")
                    return

            elif self.user['role'] == 'преподаватель':
                query = '''
                    SELECT 
                        schedule.day, 
                        schedule.date, 
                        bell_schedule.start_time, 
                        bell_schedule.end_time, 
                        subjects.subject_name, 
                        teachers.name, 
                        classrooms.classroom_name
                    FROM schedule
                    JOIN bell_schedule ON schedule.pair_number = bell_schedule.pair_number
                    JOIN subjects ON schedule.subject_id = subjects.id
                    JOIN teachers ON schedule.teacher_id = teachers.id
                    JOIN classrooms ON schedule.classroom_id = classrooms.id
                    WHERE schedule.teacher_id = ?
                    ORDER BY schedule.day, schedule.pair_number;
                '''
                cursor.execute(query, (self.user['teacher_id'],))
                rows = cursor.fetchall()

                if not rows:
                    self.label.setText("No schedule found for your teaching assignments.")
                    return

            options = QFileDialog.Options()
            file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", "",
                                                               "Excel Files (*.xlsx);;All Files (*)", options=options)

            if not file_path:
                QMessageBox.information(self, "Отмена", "Экспорт отменен.")
                return

            # Создаем новый Excel файл
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Weekly Schedule"

            # Заголовки
            days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
            columns = ["Время", "Предмет", "Преподаватель", "Аудитория"]

            # Добавляем заголовок для дней недели
            row_num = 1
            for day in days_of_week:
                col_num = days_of_week.index(day) * 4 + 1
                sheet.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=col_num + 3)
                sheet.cell(row=row_num, column=col_num, value=day)

            # Заполняем данные
            row_num = 2
            for day in days_of_week:
                # Фильтруем расписание по дням недели
                day_rows = [row for row in rows if row[0] == day]
                if day_rows:
                    col_num = days_of_week.index(day) * 4 + 1
                    for pair_num, row in enumerate(day_rows, start=1):
                        start_time, end_time, subject_name, teacher_name, classroom_name = row[2], row[3], row[4], row[
                            5], row[6]
                        sheet.cell(row=row_num + pair_num, column=col_num, value=f"{start_time} - {end_time}")
                        sheet.cell(row=row_num + pair_num, column=col_num + 1, value=subject_name)
                        sheet.cell(row=row_num + pair_num, column=col_num + 2, value=teacher_name)
                        sheet.cell(row=row_num + pair_num, column=col_num + 3, value=classroom_name)

            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Файл успешно сохранен в: {file_path}")

        except sqlite3.Error as e:
            print(f"Database query error: {e}")
        except Exception as e:
            print(f"Unexpected error: {e}")
        finally:
            conn.close()

    def search_schedule_by_date(self):
        """Поиск расписания по дате для студента или преподавателя."""
        self.search_layout = QVBoxLayout()

        self.date_label = QLabel("Введите дату (ГГГГ-ММ-ДД):")
        self.search_layout.addWidget(self.date_label)

        self.date_input = QDateEdit(self)
        self.date_input.setCalendarPopup(True)
        self.date_input.setDate(datetime.now().date())
        self.date_input.setDisplayFormat("yyyy-MM-dd")
        self.search_layout.addWidget(self.date_input)

        self.search_button = QPushButton('Найти')
        self.search_button.clicked.connect(self.display_schedule_for_date)
        self.search_layout.addWidget(self.search_button)

        self.search_widget = QWidget(self)
        self.search_widget.setLayout(self.search_layout)
        self.setCentralWidget(self.search_widget)

        self.back_button = QPushButton('Вернуться в главное меню')
        self.back_button.clicked.connect(self.back_to_main_menu)
        self.search_layout.addWidget(self.back_button)

    def display_schedule_for_date(self):
        """Отображение расписания по выбранной дате."""
        date_str = self.date_input.text()
        # Проверяем корректность даты
        try:
            search_date = QDate.fromString(date_str, "yyyy-MM-dd")
            if not search_date.isValid():
                raise ValueError("Неправильный формат даты.")
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Неправильный формат даты. Используйте ГГГГ-ММ-ДД.")
            return

        # Очистка предыдущей таблицы, если существует
        if hasattr(self, 'schedule_table') and self.schedule_table is not None:
            self.schedule_table.setParent(None)
            del self.schedule_table

        conn = connect_db()
        if conn is None:
            QMessageBox.warning(self, "Ошибка", "Ошибка подключения к базе данных.")
            return

        cursor = conn.cursor()
        try:
            if self.user['role'] == 'студент':
                query = '''
                    SELECT 
                        bell_schedule.start_time, 
                        bell_schedule.end_time, 
                        subjects.subject_name, 
                        teachers.name, 
                        classrooms.classroom_name
                    FROM schedule
                    JOIN bell_schedule ON schedule.pair_number = bell_schedule.pair_number
                    JOIN subjects ON schedule.subject_id = subjects.id
                    JOIN teachers ON schedule.teacher_id = teachers.id
                    JOIN classrooms ON schedule.classroom_id = classrooms.id
                    WHERE schedule.date = ? AND schedule.group_id = (
                        SELECT group_id FROM students WHERE id = ?
                    );
                '''
                cursor.execute(query, (search_date.toString("yyyy-MM-dd"), self.user['student_id']))
            elif self.user['role'] == 'преподаватель':
                query = '''
                    SELECT 
                        bell_schedule.start_time, 
                        bell_schedule.end_time, 
                        subjects.subject_name, 
                        classrooms.classroom_name, 
                        groups.group_name
                    FROM schedule
                    JOIN bell_schedule ON schedule.pair_number = bell_schedule.pair_number
                    JOIN subjects ON schedule.subject_id = subjects.id
                    JOIN classrooms ON schedule.classroom_id = classrooms.id
                    JOIN groups ON schedule.group_id = groups.id
                    WHERE schedule.date = ? AND schedule.teacher_id = ?;
                '''
                cursor.execute(query, (search_date.toString("yyyy-MM-dd"), self.user['teacher_id']))
            else:
                QMessageBox.warning(self, "Ошибка","Неизвестная роль пользователя.")
                return

            rows = cursor.fetchall()
            if not rows:
                QMessageBox.warning(self, "Ошибка","Расписание отсутствует на указанную дату.")
                return

            # Создание таблицы для отображения
            self.schedule_table = QTableWidget(self)
            self.schedule_table.setRowCount(len(rows))
            if self.user['role'] == 'студент':
                self.schedule_table.setColumnCount(5)
                self.schedule_table.setHorizontalHeaderLabels(
                    ['Начало', 'Конец', 'Предмет', 'Преподаватель', 'Кабинет'])
            elif self.user['role'] == 'преподаватель':
                self.schedule_table.setColumnCount(5)
                self.schedule_table.setHorizontalHeaderLabels(['Начало', 'Конец', 'Предмет', 'Кабинет', 'Группа'])

            # Заполнение таблицы
            for i, row in enumerate(rows):
                for j, value in enumerate(row):
                    self.schedule_table.setItem(i, j, QTableWidgetItem(str(value)))

            self.schedule_table.resizeColumnsToContents()
            self.schedule_table.resizeRowsToContents()
            self.search_layout.addWidget(self.schedule_table)

        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", "Ошибка базы данных: {e}")
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", "Неожиданная ошибка: {e}")
        finally:
            conn.close()

    def logout(self):
        """Функция для выхода в окно авторизации."""
        self.login_window.login_input.clear()
        self.login_window.password_input.clear()

        self.close()
        self.login_window.show()

    def back_to_main_menu(self):
        """Функция для возврата в главное меню пользователя."""
        self.close()
        self.main_window = MainWindow(self.user, self.login_window)
        self.main_window.show()

class AdminWindow(QMainWindow):
    def __init__(self, user, login_window):
        super().__init__()
        self.user = user
        self.login_window = login_window
        self.setWindowTitle('Панель администратора')
        self.setGeometry(600, 200, 800, 600)

        self.layout = QVBoxLayout()

        self.label = QLabel(f'Добро пожаловать, {user["name"]}!')
        self.layout.addWidget(self.label)

        self.add_schedule_button = QPushButton("Добавить занятие")
        self.add_schedule_button.clicked.connect(self.add_schedule)
        self.layout.addWidget(self.add_schedule_button)

        self.edit_schedule_button = QPushButton("Удалить/заменить занятия")
        self.edit_schedule_button.clicked.connect(self.edit_schedule)
        self.layout.addWidget(self.edit_schedule_button)

        self.add_user_button = QPushButton("Добавить пользователя", self)
        self.add_user_button.clicked.connect(self.add_user)
        self.layout.addWidget(self.add_user_button)

        self.logout_button = QPushButton('Выйти из аккаунта')
        self.logout_button.clicked.connect(self.logout)
        self.layout.addWidget(self.logout_button)

        main_widget = QWidget(self)
        main_widget.setLayout(self.layout)
        self.setCentralWidget(main_widget)

    def add_user(self):
        """Окно добавления нового пользователя."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавление нового пользователя")
        dialog.setModal(True)

        role_label = QLabel("Роль пользователя:")
        role_combo = QComboBox()
        role_combo.addItems(["студент", "преподаватель"])

        login_label = QLabel("Логин:")
        login_input = QLineEdit()

        password_label = QLabel("Пароль:")
        password_input = QLineEdit()
        password_input.setEchoMode(QLineEdit.Password)

        name_label = QLabel("ФИО:")
        name_input = QLineEdit()

        extra_label = QLabel("Идентификатор группы (для студента):")
        extra_combo = QComboBox()
        extra_input = QLineEdit()

        extra_input.hide()

        add_button = QPushButton("Добавить")
        cancel_button = QPushButton("Отмена")

        # Сетка компоновки
        layout = QGridLayout()
        layout.addWidget(role_label, 0, 0)
        layout.addWidget(role_combo, 0, 1)
        layout.addWidget(login_label, 1, 0)
        layout.addWidget(login_input, 1, 1)
        layout.addWidget(password_label, 2, 0)
        layout.addWidget(password_input, 2, 1)
        layout.addWidget(name_label, 3, 0)
        layout.addWidget(name_input, 3, 1)
        layout.addWidget(extra_label, 4, 0)
        layout.addWidget(extra_combo, 4, 1)
        layout.addWidget(extra_input, 4, 1)
        layout.addWidget(add_button, 5, 0)
        layout.addWidget(cancel_button, 5, 1)

        dialog.setLayout(layout)

        # Функция обновления интерфейса
        def update_fields():
            role = role_combo.currentText()
            if role == "студент":
                extra_label.setText("Идентификатор группы (для студента):")
                extra_combo.show()
                extra_input.hide()
                # Заполняем список групп из базы
                conn = connect_db()
                if conn is not None:
                    cursor = conn.cursor()
                    cursor.execute("SELECT id, group_name FROM groups")
                    groups = cursor.fetchall()
                    extra_combo.clear()
                    for group_id, group_name in groups:
                        extra_combo.addItem(f"{group_name} (ID: {group_id})", group_id)
                    conn.close()
            elif role == "преподаватель":
                extra_label.setText("Должность:")
                extra_combo.hide()
                extra_input.show()

        role_combo.currentIndexChanged.connect(update_fields)
        update_fields()

        # Логика кнопок
        def save_user():
            role = role_combo.currentText()
            login = login_input.text().strip()
            password = password_input.text().strip()
            name = name_input.text().strip()
            extra_info = extra_input.text().strip() if role == "преподаватель" else extra_combo.currentData()

            # Проверка на заполнение полей
            if not login or not password or not name or (not extra_info and role == "студент"):
                QMessageBox.warning(dialog, "Ошибка", "Все поля должны быть заполнены.")
                return

            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return

            cursor = conn.cursor()

            try:
                if role == "студент":
                    # Создаем запись студента
                    cursor.execute("INSERT INTO students (name, group_id) VALUES (?, ?)", (name, extra_info))
                    student_id = cursor.lastrowid
                    cursor.execute("INSERT INTO users (login, password, role, student_id) VALUES (?, ?, ?, ?)",
                                   (login, password, role, student_id))
                elif role == "преподаватель":
                    # Создаем запись преподавателя
                    cursor.execute("INSERT INTO teachers (name, position) VALUES (?, ?)", (name, extra_info))
                    teacher_id = cursor.lastrowid
                    cursor.execute("INSERT INTO users (login, password, role, teacher_id) VALUES (?, ?, ?, ?)",
                                   (login, password, role, teacher_id))

                conn.commit()
                QMessageBox.information(dialog, "Успех", f"Пользователь {name} успешно добавлен.")
                dialog.accept()
            except sqlite3.Error as e:
                QMessageBox.warning(dialog, "Ошибка", f"Ошибка базы данных: {e}")
            finally:
                conn.close()

        add_button.clicked.connect(save_user)
        cancel_button.clicked.connect(dialog.reject)

        dialog.exec_()

    def edit_schedule(self):
        """Окно редактирования расписания."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Изменение расписания")
        dialog.setModal(True)

        group_label = QLabel("Выберите группу:")
        group_combobox = QComboBox()

        date_label = QLabel("Выберите дату:")
        date_combobox = QComboBox()

        pair_label = QLabel("Выберите пару:")
        pair_combobox = QComboBox()

        teacher_label = QLabel("Выберите преподавателя:")
        teacher_combobox = QComboBox()

        classroom_label = QLabel("Выберите аудиторию:")
        classroom_combobox = QComboBox()

        info_label = QLabel("Информация о текущей паре:")

        save_button = QPushButton("Сохранить изменения")
        delete_button = QPushButton("Удалить пару")
        cancel_button = QPushButton("Отмена")

        layout = QGridLayout()
        layout.addWidget(group_label, 0, 0)
        layout.addWidget(group_combobox, 0, 1)
        layout.addWidget(date_label, 1, 0)
        layout.addWidget(date_combobox, 1, 1)
        layout.addWidget(pair_label, 2, 0)
        layout.addWidget(pair_combobox, 2, 1)
        layout.addWidget(info_label, 3, 0, 1, 2)
        layout.addWidget(teacher_label, 4, 0)
        layout.addWidget(teacher_combobox, 4, 1)
        layout.addWidget(classroom_label, 5, 0)
        layout.addWidget(classroom_combobox, 5, 1)
        layout.addWidget(save_button, 6, 0)
        layout.addWidget(delete_button, 6, 1)
        layout.addWidget(cancel_button, 7, 0, 1, 2)

        dialog.setLayout(layout)

        # Загрузка списка групп
        def load_groups():
            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return
            cursor = conn.cursor()
            cursor.execute("SELECT id, group_name FROM groups")
            groups = cursor.fetchall()
            group_combobox.clear()
            group_combobox.addItem("Выберите группу", -1)
            for group_id, group_name in groups:
                group_combobox.addItem(group_name, group_id)
            conn.close()

        # Загрузка дат для выбранной группы
        def load_dates():
            group_id = group_combobox.currentData()
            if group_id == -1:
                return
            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT date FROM schedule WHERE group_id = ? ORDER BY date", (group_id,))
            dates = cursor.fetchall()
            date_combobox.clear()
            date_combobox.addItem("Выберите дату", -1)
            for date, in dates:
                date_combobox.addItem(date, date)
            conn.close()

        # Загрузка пар для выбранной группы и даты
        def load_pairs():
            date = date_combobox.currentData()
            group_id = group_combobox.currentData()
            if date == -1 or group_id == -1:
                return
            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return
            cursor = conn.cursor()
            cursor.execute(
                "SELECT pair_number FROM schedule WHERE group_id = ? AND date = ? ORDER BY pair_number",
                (group_id, date)
            )
            pairs = cursor.fetchall()
            pair_combobox.clear()
            pair_combobox.addItem("Выберите пару", -1)
            for pair, in pairs:
                pair_combobox.addItem(str(pair), pair)
            conn.close()

        # Загрузка информации о выбранной паре
        def load_pair_info():
            pair_number = pair_combobox.currentData()
            group_id = group_combobox.currentData()
            date = date_combobox.currentData()
            if pair_number == -1 or group_id == -1 or date == -1:
                return
            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return
            cursor = conn.cursor()
            cursor.execute(
                '''
                SELECT 
                    subjects.subject_name, 
                    teachers.name AS teacher, 
                    classrooms.classroom_name 
                FROM schedule
                JOIN subjects ON schedule.subject_id = subjects.id
                JOIN teachers ON schedule.teacher_id = teachers.id
                JOIN classrooms ON schedule.classroom_id = classrooms.id
                WHERE schedule.group_id = ? AND schedule.date = ? AND schedule.pair_number = ?
                ''',
                (group_id, date, pair_number)
            )
            row = cursor.fetchone()
            if row:
                info_label.setText(f"<b>Текущая информация:</b><br>"
                                   f"Предмет: {row[0]}<br>"
                                   f"Преподаватель: {row[1]}<br>"
                                   f"Аудитория: {row[2]}")
                load_teachers()
                load_classrooms()
                teacher_combobox.setCurrentText(row[1])
                classroom_combobox.setCurrentText(row[2])
            conn.close()

        # Загрузка списка преподавателей
        def load_teachers():
            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM teachers")
            teachers = cursor.fetchall()
            teacher_combobox.clear()
            for teacher_id, teacher_name in teachers:
                teacher_combobox.addItem(teacher_name, teacher_id)
            conn.close()

        # Загрузка списка аудиторий
        def load_classrooms():
            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return
            cursor = conn.cursor()
            cursor.execute("SELECT id, classroom_name FROM classrooms")
            classrooms = cursor.fetchall()
            classroom_combobox.clear()
            for classroom_id, classroom_name in classrooms:
                classroom_combobox.addItem(classroom_name, classroom_id)
            conn.close()

        # Сохранение изменений
        def save_changes():
            teacher_id = teacher_combobox.currentData()
            classroom_id = classroom_combobox.currentData()
            pair_number = pair_combobox.currentData()
            group_id = group_combobox.currentData()
            date = date_combobox.currentData()

            if teacher_id == -1 or classroom_id == -1:
                QMessageBox.warning(dialog, "Ошибка", "Необходимо выбрать преподавателя и аудиторию.")
                return

            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return
            cursor = conn.cursor()

            try:
                # Проверка занятости аудитории
                cursor.execute(
                    '''
                    SELECT id FROM schedule 
                    WHERE classroom_id = ? AND pair_number = ? AND date = ? AND group_id != ?
                    ''',
                    (classroom_id, pair_number, date, group_id)
                )
                if cursor.fetchone():
                    raise ValueError("Аудитория занята в указанное время.")

                # Обновление расписания
                cursor.execute(
                    '''
                    UPDATE schedule SET teacher_id = ?, classroom_id = ? 
                    WHERE group_id = ? AND date = ? AND pair_number = ?
                    ''',
                    (teacher_id, classroom_id, group_id, date, pair_number)
                )
                conn.commit()
                QMessageBox.information(dialog, "Успех", "Изменения успешно сохранены.")
                dialog.accept()
            except ValueError as e:
                QMessageBox.warning(dialog, "Ошибка", str(e))
            except sqlite3.Error as e:
                QMessageBox.warning(dialog, "Ошибка базы данных", str(e))
            finally:
                conn.close()

        # Удаление пары
        def delete_pair():
            pair_number = pair_combobox.currentData()
            group_id = group_combobox.currentData()
            date = date_combobox.currentData()
            if pair_number == -1 or group_id == -1 or date == -1:
                QMessageBox.warning(dialog, "Ошибка", "Не выбрана пара для удаления.")
                return

            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return
            cursor = conn.cursor()

            try:
                cursor.execute(
                    '''
                    DELETE FROM schedule WHERE group_id = ? AND date = ? AND pair_number = ?
                    ''',
                    (group_id, date, pair_number)
                )
                conn.commit()
                QMessageBox.information(dialog, "Успех", "Пара успешно удалена.")
                dialog.accept()
            except sqlite3.Error as e:
                QMessageBox.warning(dialog, "Ошибка базы данных", str(e))
            finally:
                conn.close()

        # Подключение сигналов
        group_combobox.currentIndexChanged.connect(load_dates)
        date_combobox.currentIndexChanged.connect(load_pairs)
        pair_combobox.currentIndexChanged.connect(load_pair_info)

        load_groups()

        save_button.clicked.connect(save_changes)
        delete_button.clicked.connect(delete_pair)
        cancel_button.clicked.connect(dialog.reject)

        dialog.exec_()

    def add_schedule(self):
        """Окно создания расписания."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавление занятия")
        dialog.setModal(True)

        # Поля ввода
        group_label = QLabel("Группа:")
        group_combo = QComboBox()

        teacher_label = QLabel("Преподаватель:")
        teacher_combo = QComboBox()

        subject_label = QLabel("Предмет:")
        subject_combo = QComboBox()

        classroom_label = QLabel("Аудитория:")
        classroom_combo = QComboBox()

        pair_label = QLabel("Номер пары:")
        pair_combo = QComboBox()

        day_label = QLabel("День недели:")
        day_combo = QComboBox()
        day_combo.addItems(["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"])

        date_label = QLabel("Дата занятия (дд.мм.гггг):")
        date_input = QDateEdit()
        date_input.setDisplayFormat("dd.MM.yyyy")
        date_input.setCalendarPopup(True)
        date_input.setDate(datetime.now().date())

        # Кнопки
        add_button = QPushButton("Добавить")
        cancel_button = QPushButton("Отмена")

        # Сетка компоновки
        layout = QGridLayout()
        layout.addWidget(group_label, 0, 0)
        layout.addWidget(group_combo, 0, 1)
        layout.addWidget(teacher_label, 1, 0)
        layout.addWidget(teacher_combo, 1, 1)
        layout.addWidget(subject_label, 2, 0)
        layout.addWidget(subject_combo, 2, 1)
        layout.addWidget(classroom_label, 3, 0)
        layout.addWidget(classroom_combo, 3, 1)
        layout.addWidget(pair_label, 4, 0)
        layout.addWidget(pair_combo, 4, 1)
        layout.addWidget(day_label, 5, 0)
        layout.addWidget(day_combo, 5, 1)
        layout.addWidget(date_label, 6, 0)
        layout.addWidget(date_input, 6, 1)
        layout.addWidget(add_button, 7, 0)
        layout.addWidget(cancel_button, 7, 1)

        dialog.setLayout(layout)

        # Заполнение данных
        def load_data():
            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return

            cursor = conn.cursor()

            try:
                cursor.execute("SELECT id, group_name FROM groups")
                for group_id, group_name in cursor.fetchall():
                    group_combo.addItem(f"{group_name} (ID: {group_id})", group_id)

                cursor.execute("SELECT id, name FROM teachers")
                for teacher_id, teacher_name in cursor.fetchall():
                    teacher_combo.addItem(f"{teacher_name} (ID: {teacher_id})", teacher_id)

                cursor.execute("SELECT id, classroom_name FROM classrooms")
                for classroom_id, classroom_name in cursor.fetchall():
                    classroom_combo.addItem(f"№{classroom_name} (ID: {classroom_id})", classroom_id)

                cursor.execute("SELECT pair_number, start_time, end_time FROM bell_schedule")
                for pair_number, start_time, end_time in cursor.fetchall():
                    pair_combo.addItem(f"{pair_number} - {start_time}–{end_time}", pair_number)
            except sqlite3.Error as e:
                QMessageBox.warning(dialog, "Ошибка", f"Ошибка базы данных: {e}")
            finally:
                conn.close()

        # Загрузка предметов для выбранного преподавателя
        def update_subjects():
            teacher_id = teacher_combo.currentData()
            subject_combo.clear()

            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return

            cursor = conn.cursor()
            try:
                cursor.execute("""
                    SELECT DISTINCT subjects.id, subjects.subject_name 
                    FROM subjects
                    JOIN schedule ON schedule.subject_id = subjects.id
                    WHERE schedule.teacher_id = ?
                """, (teacher_id,))
                subjects = cursor.fetchall()
                if subjects:
                    for subject_id, subject_name in subjects:
                        subject_combo.addItem(f"{subject_name} (ID: {subject_id})", subject_id)
                else:
                    QMessageBox.information(dialog, "Информация",
                                            "У выбранного преподавателя нет закрепленных предметов.")
            except sqlite3.Error as e:
                QMessageBox.warning(dialog, "Ошибка", f"Ошибка базы данных: {e}")
            finally:
                conn.close()

        # Логика кнопок
        def save_schedule():
            group_id = group_combo.currentData()
            teacher_id = teacher_combo.currentData()
            subject_id = subject_combo.currentData()
            classroom_id = classroom_combo.currentData()
            pair_id = pair_combo.currentData()
            day_of_week = day_combo.currentText()
            date_of_class = date_input.date().toString("yyyy-MM-dd")

            # Проверка на заполнение полей
            if not all([group_id, teacher_id, subject_id, classroom_id, pair_id]):
                QMessageBox.warning(dialog, "Ошибка", "Все поля должны быть заполнены.")
                return

            conn = connect_db()
            if conn is None:
                QMessageBox.warning(dialog, "Ошибка", "Ошибка подключения к базе данных.")
                return

            cursor = conn.cursor()

            try:
                # Проверка наличия пары у выбранной группы на заданный день, дату и номер пары
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM schedule 
                    WHERE group_id = ? AND date = ? AND pair_number = ? AND day = ?
                """, (group_id, date_of_class, pair_id, day_of_week))
                group_conflict = cursor.fetchone()[0]

                if group_conflict > 0:
                    QMessageBox.warning(dialog, "Ошибка",
                                        "У выбранной группы уже существует пара на указанное время.")
                    return

                # Проверка занятости аудитории
                cursor.execute("""
                    SELECT COUNT(*) 
                    FROM schedule 
                    WHERE classroom_id = ? AND date = ? AND pair_number = ? AND day = ?
                """, (classroom_id, date_of_class, pair_id, day_of_week))
                room_conflict = cursor.fetchone()[0]

                if room_conflict > 0:
                    QMessageBox.warning(dialog, "Ошибка",
                                        "Указанная аудитория уже занята на указанное время.")
                    return

                # Добавляем запись в таблицу расписания
                cursor.execute("""
                    INSERT INTO schedule (day, date, pair_number, subject_id, teacher_id, classroom_id, group_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (day_of_week, date_of_class, pair_id, subject_id, teacher_id, classroom_id, group_id))

                conn.commit()
                QMessageBox.information(dialog, "Успех", "Расписание успешно добавлено.")
                dialog.accept()
            except sqlite3.Error as e:
                QMessageBox.warning(dialog, "Ошибка", f"Ошибка базы данных: {e}")
            finally:
                conn.close()

        teacher_combo.currentIndexChanged.connect(update_subjects)
        add_button.clicked.connect(save_schedule)
        cancel_button.clicked.connect(dialog.reject)

        load_data()
        dialog.exec_()

    def back_to_main_menu(self):
        """Функция для возврата в главное меню пользователя."""
        self.close()
        self.main_window = AdminWindow(self.user, self.login_window)
        self.main_window.show()
    def logout(self):
        """Выход администратора."""
        self.login_window.login_input.clear()
        self.login_window.password_input.clear()
        self.close()
        self.login_window.show()
        
class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Авторизация')
        self.setGeometry(750, 350, 400, 300)

        layout = QVBoxLayout()

        self.login_label = QLabel('Логин:')
        layout.addWidget(self.login_label)

        self.login_input = QLineEdit(self)
        layout.addWidget(self.login_input)

        self.password_label = QLabel('Пароль:')
        layout.addWidget(self.password_label)

        self.password_input = QLineEdit(self)
        self.password_input.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.password_input)

        self.login_button = QPushButton('Войти', self)
        self.login_button.clicked.connect(self.check_credentials)
        layout.addWidget(self.login_button)

        self.setLayout(layout)

    def check_credentials(self):
        """Проверка данных пользователя для входа."""
        login = self.login_input.text()
        password = self.password_input.text()
        # Проверка на пустые поля
        if not login.strip() or not password.strip():
            QMessageBox.warning(self, "Ошибка", "Имя пользователя и пароль не могут быть пустыми.")
            return
        conn = connect_db()
        if conn is None:
            QMessageBox.warning(self, "Ошибка", "Ошибка подключения к базе данных.")
            return
        cursor = conn.cursor()
        try:
            cursor.execute('SELECT id, role, student_id, teacher_id FROM users WHERE login = ? AND password = ?',
                           (login, password))
            user = cursor.fetchone()
            if user:
                user_data = {'id': user[0], 'role': user[1], 'student_id': user[2], 'teacher_id': user[3]}
                # Получаем имя в зависимости от роли
                if user_data['role'] == 'студент':
                    cursor.execute('SELECT name FROM students WHERE id = ?', (user_data['student_id'],))
                elif user_data['role'] == 'преподаватель':
                    cursor.execute('SELECT name FROM teachers WHERE id = ?', (user_data['teacher_id'],))
                elif user_data['role'] == 'администратор':
                    user_data['name'] = "Администратор"
                else:
                    QMessageBox.warning(self, "Ошибка", "Неизвестная роль пользователя.")
                    return

                # Если имя для студента или преподавателя запрошено, то добавляем его
                if user_data['role'] in ['студент', 'преподаватель']:
                    name_result = cursor.fetchone()
                    if name_result:
                        user_data['name'] = name_result[0]
                    else:
                        QMessageBox.warning(self, "Ошибка", "Не удалось найти данные пользователя.")
                        return
                self.open_main_window(user_data)
            else:
                QMessageBox.warning(self, "Ошибка", "Неверное имя пользователя или пароль.")
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Ошибка", f"Ошибка базы данных: {e}")
        finally:
            conn.close()

    def open_main_window(self, user):
        if user['role'] == 'администратор':
            self.close()
            self.admin_window = AdminWindow(user, self)
            self.admin_window.show()
        else:
            self.close()
            self.main_window = MainWindow(user, self)
            self.main_window.show()

if __name__ == "__main__":
    app = QApplication([])
    login_window = LoginWindow()
    login_window.show()
    app.exec_()

